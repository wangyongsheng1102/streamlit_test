import gc
import os
import shutil

import openpyxl


def copy_folder(src_folder, dest_folder):
    try:
        if not os.path.exists(dest_folder):
            os.makedirs(dest_folder)

        for item in os.listdir(src_folder):
            item_path = os.path.join(src_folder, item)
            dest_path = os.path.join(dest_folder, item)

            if os.path.isdir(item_path):
                copy_folder(item_path, dest_path)
            else:
                shutil.copy2(item_path, dest_path)

        print(f"Folder '{src_folder}' copied to '{dest_folder}' successfully.")
    except Exception as e:
        print(f"An error occurred: {e}")


def del_folder(folder_path):
    try:
        shutil.rmtree(folder_path)
    except OSError as e:
        print(f"while deleting '{folder_path}' error：{e.strerror}")


def parse_excel(file_path):
    wb = openpyxl.load_workbook(file_path, data_only=False)
    try:
        ws = wb["WBS"]
    except KeyError:
        raise ValueError("WBSファイルが誤って、チェックしてください。")
    try:
        all_data = []
        for row in ws.iter_rows():
            if row[0].row < 8:
                continue
            all_data.append([row[2].value, row[3].value, row[4].value, row[5].value, ])
        qa_data = []
        for row in ws.iter_rows():
            if row[0].row < 8:
                continue
            if row[10].value is not None:
                qa_data.append([row[2].value, row[3].value, row[4].value, row[5].value, row[10].value])
    finally:
        wb.close()
        del wb
        gc.collect()

    return all_data, qa_data


def column_letter_to_number(column_letter):
    """ディジットに変更する"""
    column_number = 0
    for char in column_letter:
        column_number = column_number * 26 + (ord(char) - ord("A") + 1)
    return column_number


def get_str_before_first_dot(string, split):
    return string.split(split)[0]


def get_str_after_last_dot(string, split):
    return string.split(split)[len(string.split(split)) - 1]
