import gc
import time
from io import BytesIO

import numpy as np
import openpyxl
import pandas as pd
import streamlit as st
from matplotlib import pyplot as plt
from matplotlib.animation import FuncAnimation

from utils.layout import add_separator_rainbow

from utils.logging import log_message

WBS_FILE_PATH = "D:/BIP/3開発庫/01連絡/01スケジュール/【内部】【Enability_Spring化対応】WBS管理台帳.xlsm"
WBS_SHEET_NAME = "WBS"


def parse_excel(file_path):
    wb = openpyxl.load_workbook(file_path, data_only=True)
    try:
        ws = wb["WBS"]
    except KeyError:
        raise ValueError("WBSファイルが誤って、チェックしてください。")
    try:
        all_data = []
        all_row = []
        for row in ws.iter_rows():
            if row[0].row < 8:
                continue
            if row[18].value == "作業完了":
                continue
            show_flag = False
            for cell in row:
                if isinstance(cell.value, str) and cell.value == st.session_state['name']:
                    show_flag = True
            if show_flag:
                all_data.append([row[2].value, row[3].value, row[4].value, row[5].value, ])
                temp_list = []
                for cell in row:
                    temp_list.append(cell.value)
                all_row.append(temp_list)
        # qa_data = []
        # for row in ws.iter_rows():
        #     if row[0].row < 8:
        #         continue
        #     if row[10].value is not None:
        #         qa_data.append([row[2].value, row[3].value, row[4].value, row[5].value, row[10].value])
    finally:
        wb.close()
        del wb
        gc.collect()

    return all_data, all_row


def render():
    st.title(":school_satchel: :gray[_任務一覧_] :school_satchel:")
    add_separator_rainbow()

    log_message(f"ページ「任務一覧」は表示しました。", "INFO")

    with st.spinner('WBSファイルは読み込み中...'):
        if 'work_list' in st.session_state and st.session_state['work_list'] is not None:
            st.text('残作業一覧')
            st.dataframe(st.session_state['work_list'])

            st.text('for test')
            st.dataframe(st.session_state['work_list_all'])
        else:
            df, df_all = parse_excel(WBS_FILE_PATH)
            # qa_df = pd.DataFrame(qa)
            # qa_df.columns = ["対象区分", "システム", "機能ID", "機能名", "備考"]
            # st.text('QA残')
            # # st.markdown("--------------------------------------------")
            # st.dataframe(qa_df)
            st.text('残作業一覧')
            df_df = pd.DataFrame(df)
            df_df.columns = ["対象区分", "システム", "機能ID", "機能名"]
            st.dataframe(df_df)
            st.session_state['work_list'] = df_df

            st.text('for test')
            df_all_df = pd.DataFrame(df_all)
            # df_all_df.columns = ["対象区分", "システム", "機能ID", "機能名"]
            st.dataframe(df_all_df)
            st.session_state['work_list_all'] = df_all_df
